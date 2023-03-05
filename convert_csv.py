import csv
import openpyxl
from openpyxl.styles import Border,Side,PatternFill
import pathlib
#*** insert your text file path here ***
text_file=r"C:\Users\Design\Desktop\AE\RTES\logs\sample-28022023.txt"
#***************************************
header=['Total Fuel Pulses','Total Water Pulses','Fuel Pulse Period (s)','Fuel Flowrate (mL/min)','Cycle Position (/5)','Water Percentage (%)','Fuel Percentage (%)']
text_file.replace('\\','/')
parent_dir=pathlib.PurePath(text_file).parent
file=pathlib.PurePath(text_file).stem
xlsx_file=parent_dir.joinpath(file+'.xlsx')
fuel_bias=1.45
water_bias=0.81
#cycles
cycle_column=['E:E',4]
cycle_count=5
total_cycles=0
#fuel pulse
fuel_pulse_column='A:A'
fuel_flow_rate_column='D:D'
fuel_pulse_period_column='C:C'
#water pulse
water_pulse_column='B:B'
wb=openpyxl.Workbook()
ws=wb.worksheets[0]
with open(text_file, 'r') as txt_file:
    text_file=filter(lambda txt: len(txt)==7 or len(txt)==8 or (txt[0]!='[' and txt.find(',') != -1), txt_file)
    reader=csv.reader(text_file,delimiter=',')
    ws.append(header)
    for count,row in enumerate(reader):
        row=[float(i) for i in row] if str(row).find(',') != -1 else row
        ws.append(row)
        try:
            if row[cycle_column[1]]==cycle_count:
                total_cycles+=1
        except:
            pass
    print(total_cycles)
    ws['J1'].value='Total Cycles'
    ws['J1'].border=Border(left=Side(border_style='thin'),
                           right=Side(border_style='thin'),
                           top=Side(border_style='thin'),
                           bottom=Side(border_style='thin'))
    ws['J1'].fill=PatternFill(start_color='0000FF00',end_color='0000FF00',fill_type='solid')

    ws['K1'].value=total_cycles
    ws['K1'].border=Border(left=Side(border_style='thin'),
                           right=Side(border_style='thin'),
                           top=Side(border_style='thin'),
                           bottom=Side(border_style='thin'))
    ws['K1'].fill=PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')

    ws['J2'].value='Total Fuel Pulses'
    ws['J2'].border=Border(left=Side(border_style='thin'),
                           right=Side(border_style='thin'),
                           top=Side(border_style='thin'),
                           bottom=Side(border_style='thin'))
    ws['J2'].fill=PatternFill(start_color='0000FF00',end_color='0000FF00',fill_type='solid')

    ws['K2'].value=f'=MAX({fuel_pulse_column})'
    ws['K2'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['K2'].fill=PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')

    ws['J3'].value='Average  Fuel Flow Rate'
    ws['J3'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['J3'].fill=PatternFill(start_color='0000FF00',end_color='0000FF00',fill_type='solid')

    ws['K3'].value=f'=AVERAGE({fuel_flow_rate_column})'
    ws['K3'].number_format='0.00" mL/min"'
    ws['K3'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['K3'].fill=PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')

    ws['J4'].value='Total Fuel Consumption'
    ws['J4'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['J4'].fill=PatternFill(start_color='0000FF00',end_color='0000FF00',fill_type='solid')

    ws['K4'].value=f'=K2*{fuel_bias}'
    ws['K4'].number_format='0.00" mL"'
    ws['K4'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['K4'].fill=PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')

    ws['J6'].value='Total Water Pulses'
    ws['J6'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['J6'].fill=PatternFill(start_color='0000CCFF',end_color='0000CCFF',fill_type='solid')

    ws['K6'].value=f'=MAX({water_pulse_column})'
    ws['K6'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['K6'].fill=PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')

    ws['J7'].value='Total Water Consumption'
    ws['J7'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['J7'].fill=PatternFill(start_color='0000CCFF',end_color='0000CCFF',fill_type='solid')

    ws['K7'].value=f'=K6*{water_bias}'
    ws['K7'].number_format='0.00" mL"'
    ws['K7'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['K7'].fill=PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')

    ws['J5'].value='Fuel Percentage'
    ws['J5'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['J5'].fill=PatternFill(start_color='0000FF00',end_color='0000FF00',fill_type='solid')

    ws['K5'].value=f'=K4/(K4+K7)*100'
    ws['K5'].number_format='0.00" %"'
    ws['K5'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['K5'].fill=PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')

    ws['J8'].value='Water Percentage'
    ws['J8'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['J8'].fill=PatternFill(start_color='0000CCFF',end_color='0000CCFF',fill_type='solid')
    
    ws['K8'].value=f'=K7/(K4+K7)*100'
    ws['K8'].number_format='0.00" %"'
    ws['K8'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['K8'].fill=PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')

    ws['J9'].value='Time Spent'
    ws['J9'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['J9'].fill=PatternFill(start_color='00FF6600',end_color='00FF6600',fill_type='solid')

    ws['K9'].value=f'=FLOOR(SUM({fuel_pulse_period_column})/60/60,1)&" h "&ROUND(SUM({fuel_pulse_period_column})/60/60-FLOOR(SUM({fuel_pulse_period_column})/60/60,1),1)*60&" m "'
    ws['K9'].border=Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
    ws['K9'].fill=PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
wb.save(xlsx_file)